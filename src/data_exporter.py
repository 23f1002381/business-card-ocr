"""
Data Export Module

This module handles exporting extracted business card data to Excel format.
"""
import os
import logging
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Union, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Initialize logger
logger = logging.getLogger(__name__)

class BusinessCardExporter:
    """
    Handles exporting business card data to Excel format.
    """
    
    # Default column widths
    COLUMN_WIDTHS = {
        'name': 25,
        'job_title': 30,
        'company': 30,
        'email': 35,
        'phone': 20,
        'website': 30,
        'address': 50,
        'notes': 40,
        'extraction_date': 20,
    }
    
    # Style definitions
    HEADER_STYLE = {
        'font': Font(bold=True, color='FFFFFF'),
        'fill': PatternFill("solid", fgColor='4F81BD'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    DATA_STYLE = {
        'alignment': Alignment(vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }
    
    def __init__(self, output_dir: Optional[Union[str, Path]] = None):
        """
        Initialize the exporter.
        
        Args:
            output_dir: Directory to save exported files. Defaults to 'exports' in current directory.
        """
        self.output_dir = Path(output_dir) if output_dir else Path.cwd() / 'exports'
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def export_to_excel(
        self,
        cards_data: List[Dict[str, Any]],
        filename: Optional[str] = None,
        sheet_name: str = 'Business Cards',
        include_timestamp: bool = True
    ) -> str:
        """
        Export business card data to an Excel file.
        
        Args:
            cards_data: List of dictionaries containing business card data
            filename: Output filename (without extension). If None, generates a name.
            sheet_name: Name of the Excel sheet
            include_timestamp: Whether to include a timestamp in the filename
            
        Returns:
            Path to the generated Excel file
        """
        try:
            if not cards_data:
                raise ValueError("No business card data provided")
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                base_name = f"business_cards_{timestamp}" if include_timestamp else "business_cards"
                filename = f"{base_name}.xlsx"
            elif not filename.endswith('.xlsx'):
                filename = f"{filename}.xlsx"
            
            output_path = self.output_dir / filename
            
            # Create a DataFrame from the data
            df = self._prepare_dataframe(cards_data)
            
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Get workbook and worksheet objects
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]
                
                # Apply formatting
                self._format_worksheet(worksheet, df)
                
                # Adjust column widths
                self._adjust_column_widths(worksheet, df)
                
                # Freeze header row
                worksheet.freeze_panes = 'A2'
                
                # Save the workbook
                workbook.save(output_path)
            
            logger.info(f"Successfully exported {len(cards_data)} business cards to {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error exporting to Excel: {str(e)}")
            raise
    
    def _prepare_dataframe(self, cards_data: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Prepare the data for export by normalizing and structuring it.
        
        Args:
            cards_data: List of business card data dictionaries
            
        Returns:
            Formatted pandas DataFrame
        """
        # Create a copy to avoid modifying the original data
        data = []
        
        for card in cards_data:
            # Create a flat dictionary for the card
            card_data = {
                'name': card.get('name', ''),
                'job_title': card.get('job_title', ''),
                'company': card.get('company', ''),
                'email': ', '.join(card.get('emails', [])),
                'phone': ', '.join(card.get('phones', [])),
                'website': ', '.join(card.get('urls', [])),
                'address': card.get('address', ''),
                'notes': card.get('notes', ''),
                'extraction_date': card.get('extraction_date', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                'source_image': card.get('source_image', '')
            }
            
            # Add any additional fields
            for key, value in card.items():
                if key not in card_data and key not in ['emails', 'phones', 'urls']:
                    if isinstance(value, (list, tuple)):
                        card_data[key] = ', '.join(str(v) for v in value)
                    else:
                        card_data[key] = str(value) if value is not None else ''
            
            data.append(card_data)
        
        # Create DataFrame with consistent column order
        columns = [
            'name', 'job_title', 'company', 'email', 'phone', 
            'website', 'address', 'notes', 'extraction_date', 'source_image'
        ]
        
        # Add any additional columns
        for card in data:
            for key in card:
                if key not in columns:
                    columns.append(key)
        
        # Create DataFrame with consistent columns
        df = pd.DataFrame(data, columns=columns)
        
        # Fill NaN values with empty strings
        df = df.fillna('')
        
        return df
    
    def _format_worksheet(self, worksheet, df):
        """Apply formatting to the worksheet."""
        # Apply header style
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.font = self.HEADER_STYLE['font']
            cell.fill = self.HEADER_STYLE['fill']
            cell.alignment = self.HEADER_STYLE['alignment']
            cell.border = self.HEADER_STYLE['border']
        
        # Apply data style
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, 
                                     min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.alignment = self.DATA_STYLE['alignment']
                cell.border = self.DATA_STYLE['border']
        
        # Auto-filter
        worksheet.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{worksheet.max_row}"
    
    def _adjust_column_widths(self, worksheet, df):
        """Adjust column widths based on content."""
        for idx, column in enumerate(df.columns, 1):
            # Get the max width from content or use default
            max_length = max(
                df[column].astype(str).apply(len).max(),  # Max length in column
                len(str(column))  # Length of column name
            )
            
            # Use predefined width if available, otherwise use content-based width
            col_letter = get_column_letter(idx)
            default_width = self.COLUMN_WIDTHS.get(column.lower(), 15)
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            final_width = max(default_width, adjusted_width)
            
            worksheet.column_dimensions[col_letter].width = final_width
    
    def export_multiple_sheets(
        self,
        sheets_data: Dict[str, List[Dict[str, Any]]],
        filename: Optional[str] = None,
        include_timestamp: bool = True
    ) -> str:
        """
        Export multiple sheets of business card data to a single Excel file.
        
        Args:
            sheets_data: Dictionary with sheet names as keys and card data as values
            filename: Output filename (without extension)
            include_timestamp: Whether to include a timestamp in the filename
            
        Returns:
            Path to the generated Excel file
        """
        try:
            if not sheets_data:
                raise ValueError("No sheet data provided")
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                base_name = f"business_cards_{timestamp}" if include_timestamp else "business_cards"
                filename = f"{base_name}.xlsx"
            elif not filename.endswith('.xlsx'):
                filename = f"{filename}.xlsx"
            
            output_path = self.output_dir / filename
            
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, cards in sheets_data.items():
                    if not cards:
                        continue
                        
                    # Prepare data for this sheet
                    df = self._prepare_dataframe(cards)
                    
                    # Write to Excel
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Get worksheet and apply formatting
                    worksheet = writer.sheets[sheet_name]
                    self._format_worksheet(worksheet, df)
                    self._adjust_column_widths(worksheet, df)
                    worksheet.freeze_panes = 'A2'
                
                # Save the workbook
                writer.book.save(output_path)
            
            total_cards = sum(len(cards) for cards in sheets_data.values())
            logger.info(f"Successfully exported {total_cards} business cards to {output_path} across {len(sheets_data)} sheets")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error exporting multiple sheets to Excel: {str(e)}")
            raise


def create_exporter(output_dir: Optional[Union[str, Path]] = None) -> BusinessCardExporter:
    """
    Create a BusinessCardExporter instance.
    
    Args:
        output_dir: Directory to save exported files
        
    Returns:
        BusinessCardExporter instance
    """
    return BusinessCardExporter(output_dir=output_dir)
